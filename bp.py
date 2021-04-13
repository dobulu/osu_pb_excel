import os
import json
import requests
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
 
class Osu:
    def __init__(self, id):
        self.url = 'https://osu.ppy.sh/api/'
        self.key = 'KEY' #Change this to your apiV1 KEY
        self.id = id

    def get_mods(self, mods_int):
        mods = ['NF', 'EZ', 'TD', 'HD', 'HR', 'SD', 'DT', 'RX', 'HT', 
                'NC', 'FL', 'AU', 'SO', 'RX2', 'PF']
        check = list(bin(mods_int).lstrip('0b'))
        check.reverse()
        mods_list = []
        mods_str = ""
        count = 0
        if mods_int == 0:
            mods_str += 'None'
        else:
            for num in check:
                if num == '1':
                    mods_list.append(mods[count])
                count += 1

            if 'NC' in mods_list:
                mods_list.remove('DT')
            if 'PF' in mods_list:
                mods_list.remove('SD')

        for mod in mods_list:
            mods_str += mod   
        return mods_str

    def get_length(self, sec):
        m, s = divmod(sec, 60)
        length = "%02d分%02d秒" % (m, s)
        return length

    def get_acc(self, n300, n100, n50):
        accuracy = 100* (50*n50 + 100*n100 + 300*n300) / (300*(n50 + n100 + n300))
        acc = f"{round(accuracy, 2)}%"
        return acc

    def get_data(self, func, PARAMS):
        r = requests.get(url = (self.url + func), params = PARAMS) 
        json_str = r.content.decode()
        temp_dict = json.loads(json_str)
        return temp_dict
    
    def get_user(self, mode=0):
        PARAMS = {'k':self.key, 'u':self.id, 'm':mode}  
        user_info = self.get_data('get_user', PARAMS)[0]
        return user_info

    def get_user_image(self):
        url='https://a.ppy.sh/' + self.get_user()['user_id']
        res=requests.get(url)
        img=res.content
        with open( './img.jpg','wb' ) as f:
            f.write(img)

    def get_user_pb(self, mode=0, limit=100):
        PARAMS = {'k':self.key, 'u':self.id, 'm':mode, 'limit': limit}  
        user_pb = self.get_data('get_user_best', PARAMS)
        return user_pb

    def get_beatmaps(self, sid):
        PARAMS = {'k':self.key, 'b':sid}  
        beatmap_info = self.get_data('get_beatmaps', PARAMS)
        return beatmap_info

    def get_beatmap(self, sid):
        beatmap_info = self.get_beatmaps(sid)
        diff_info = []
        for temp in beatmap_info:
            if int(temp['beatmap_id']) == int(sid):
                diff_info = temp
                break
        return diff_info

    def get_match(self, id):
        PARAMS = {'k':self.key, 'mp':id}  
        match_info = self.get_data('get_match', PARAMS)
        return match_info

    def excel_initial(self, ws):
        user_info = self.get_user()
        ws.cell(row=1, column=2, value= 'Name')
        ws.cell(row=2, column=2, value= 'pp')
        ws.cell(row=3, column=2, value= 'pc')
        ws.cell(row=4, column=2, value= 'acc')
        ws.cell(row=5, column=2, value= 'rank')

        link = 'https://osu.ppy.sh/users/' + user_info['user_id']
        name = user_info['username']
        ws.cell(row=1, column=3, value=('=HYPERLINK("%s","%s")' % (link, name)))
        ws.cell(row=2, column=3, value= user_info['pp_raw'])
        ws.cell(row=3, column=3, value= user_info['playcount'])
        ws.cell(row=4, column=3, value= f"{round(float(user_info['accuracy']), 2)}%")
        ws.cell(row=5, column=3, value= f" #{user_info['pp_rank']}")

    def excel_data(self, ws):
        title = ['bp', 'sid', '歌曲', '作者', '时长', 'diff', 'bpm','难度','cs', 
                 'hp', 'od', 'ar', 'mods', '300', '100', '50', 'combo', 'miss', 
                 'acc', 'rank', 'pp']
        ws.append(title)

        pbs = self.get_user_pb()
        count = 0
        for pb in pbs:
            count += 1
            print(f"{count}/100")
            beatmap_info = self.get_beatmap(pb['beatmap_id'])

            bp = f"#{count}"
            sid = pb['beatmap_id']
            path = f"osu://s/{beatmap_info['beatmapset_id']}"
            song = f"{beatmap_info['artist']} - {beatmap_info['title_unicode']}"
            mapper = beatmap_info['creator']
            length =  self.get_length(int(beatmap_info['total_length']))
            diff = beatmap_info['version']
            bpm = float(beatmap_info['bpm'])
            difficulty = round(float(beatmap_info['difficultyrating']), 2)
            cs = float(beatmap_info['diff_size'])
            hp = float(beatmap_info['diff_drain'])
            od = float(beatmap_info['diff_overall'])
            ar = float(beatmap_info['diff_approach'])
            modes = self.get_mods(int(pb['enabled_mods']))
            n300 = int(pb['count300'])
            n100 = int(pb['count100'])
            n50 = int(pb['count50'])
            combo = f"{int(pb['maxcombo'])}/{int(beatmap_info['max_combo'])}"
            miss = int(pb['countmiss'])
            acc = self.get_acc(n300, n100, n50)
            rank = pb['rank']
            pp = round(float(pb['pp']), 2)

            pb_info = [bp, None, song, mapper, length, diff, bpm, difficulty, cs,
                       hp, od, ar, modes, n300, n100, n50, combo, miss, acc, 
                       rank, pp]
            ws.append(pb_info)
            ws.cell(row=(count+6), column=2, 
                    value=('=HYPERLINK("%s","%s")' % (path, sid)))

    def excel_fit_width(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    def excel_from_center(self, ws):
        nrows = ws.max_row  
        ncols = ws.max_column
        for i in range(nrows):
            for j in range(ncols):
                ws.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
    
    def excel_image(self, ws):
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        for i in range(1,6):
            ws.row_dimensions[i].height = 15
        ws.merge_cells('A1:A5')

        self.get_user_image()
        img = Image('./img.jpg')
        newSize=(100,100)
        img.width,img.height = newSize
        ws.add_image(img,'A1') 

    def excel_pb(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        self.excel_initial(ws)
        self.excel_data(ws)
        self.excel_fit_width(ws)
        self.excel_from_center(ws)
        self.excel_image(ws)
        wb.save(f"{self.id}.xlsx")
        os.remove('./img.jpg')

if __name__ == '__main__':
    while True:
        id = input("输入名称或id：")
        try:
            user = Osu(id)
            user.excel_pb()
        except IndexError:
            pass
